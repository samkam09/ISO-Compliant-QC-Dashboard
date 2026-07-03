import pandas as pd
import os
import numpy as np
import time
import re
from datetime import datetime
from tqdm import tqdm
import warnings
from openpyxl.worksheet.table import Table, TableStyleInfo
import pickle
import concurrent.futures
import win32com.client as win32
from collections import defaultdict

# Suppress openpyxl warnings for a cleaner console
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
# =====================================================================
# GLOBAL CONFIGURATION
# =====================================================================
PATH_LIST_FILE = r"S:\Ops-OPH\OPHA-OPHA\Limited\OEC-OPHA\UKAS\Ref Material checks\folder_list.xlsx"
FINAL_OUTPUT_DIR = r"S:\Ops-OPH\OPHA-OPHA\Limited\OEC-OPHA\UKAS\Ref Material checks"
if not os.path.exists(FINAL_OUTPUT_DIR):
    os.makedirs(FINAL_OUTPUT_DIR)
# =====================================================================
# UNIVERSAL HELPER FUNCTIONS
# =====================================================================
def get_folders_from_excel(column_identifier, fallback_index):
    """Dynamically locates folder paths from the master Excel list."""
    try:
        df_paths = pd.read_excel(PATH_LIST_FILE)
        df_paths.columns = df_paths.columns.str.strip()
        
        # Check by string matching in column names
        match_cols = [c for c in df_paths.columns if column_identifier.lower() in c.lower()]
        if match_cols:
            return df_paths[match_cols[0]].dropna().tolist()
        else:
            return df_paths.iloc[:, fallback_index].dropna().tolist()
    except Exception as e:
        print(f"❌ Error reading {PATH_LIST_FILE} for {column_identifier}: {e}")
        return []

def save_ukas_formatted_table(df, filename, table_name, col_a_width=15, col_b_width=15):
    """Applies official UKAS/ISO Excel Table formatting universally."""
    if df.empty:
        return
        
    output_path = os.path.join(FINAL_OUTPUT_DIR, filename)
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, header=True, sheet_name="QC_Data")
        sheet = writer.book.active

        last_row = sheet.max_row
        last_col_letter = sheet.cell(row=1, column=sheet.max_column).column_letter
        
        table = Table(displayName=table_name, ref=f"A1:{last_col_letter}{last_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=True)
        sheet.add_table(table)
        
        sheet.column_dimensions['A'].width = col_a_width
        sheet.column_dimensions['B'].width = col_b_width

        # Format numerical data to 2 decimal places (starting Col C)
        for row in sheet.iter_rows(min_row=2, min_col=3):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00'

    print(f"✅ Final Table successfully saved at: {output_path}")

def resolve_duplicate_columns(df):
    """Prevents openpyxl corruption by renaming duplicate headers."""
    new_cols, seen = [], {}
    for c in df.columns:
        c_str = str(c)
        if c_str in seen:
            seen[c_str] += 1
            new_cols.append(f"{c_str}_{seen[c_str]}")
        else:
            seen[c_str] = 0
            new_cols.append(c_str)
    df.columns = new_cols
    return df

# =====================================================================
# MODULE 1: PCBs & DIOXINS
# =====================================================================
def process_pcb_dioxins():
    print("\n" + "="*50 + "\n🚀 INITIATING PCB & DIOXIN EXTRACTION\n" + "="*50)
    pcb_folders = get_folders_from_excel("PCB", fallback_index=1)
    dioxin_folders = get_folders_from_excel("Dioxin", fallback_index=2)
    
    pcb_rows = ["PCB 28", "PCB 52", "PCB 101", "PCB 105", "PCB 118", "PCB 138", "PCB 153", "PCB 156", "PCB 180", "PCB 114", "PCB 123", "PCB 157", "PCB 167", "PCB 189"]
    dxn_rows = ["Name", "2,3,7,8-TCDD", "1,2,3,7,8-PeCDD", "1,2,3,6,7,8-HxCDD", "1,2,3,7,8,9-HxCDD", " ", "2,3,7,8-TCDF", "1,2,3,7,8-PeCDF", "2,3,4,7,8-PeCDF", "1,2,3,4,7,8-HxCDF", "1,2,3,6,7,8-HxCDF", "2,3,4,6,7,8-HxCDF", "1,2,3,4,6,7,8-HpCDF", " ", " ", "PCB 77", "PCB 126", "PCB 169", "1,2,3,4,7,8-HxCDD", "1,2,3,4,6,7,8-HpCDD", "OCDD", "1,2,3,7,8,9-HxCDF", "1,2,3,4,7,8,9-HpCDF", "OCDF", "PCB 81"]

    # --- PCB Extraction (ONLY LAST FOLDER) ---
    if pcb_folders:
        last_pcb_folder = str(pcb_folders[-1]).strip()
        out_path = os.path.join(last_pcb_folder, "consolidated_combined.xlsx")
        
        if os.path.exists(last_pcb_folder):
            qa_data, rep_data = {}, {}
            files = [f for f in os.listdir(last_pcb_folder) if f.endswith(".xlsm") and not f.startswith("~$")]
            for file in tqdm(files, desc=f"🔍 PCBs in {os.path.basename(last_pcb_folder)}"):
                header, fpath = file[:4], os.path.join(last_pcb_folder, file)
                try:
                    df_qa = pd.read_excel(fpath, sheet_name="QA", usecols="F", skiprows=2, nrows=9, engine="openpyxl")
                    qa_data[header] = [pd.read_excel(fpath, sheet_name="QA", usecols="F", nrows=1, engine="openpyxl").iloc[0, 0]] + df_qa.iloc[:, 0].tolist()
                except: pass
                try:
                    df_rep = pd.read_excel(fpath, sheet_name="Report_Regulated", usecols="F", skiprows=65, nrows=14, engine="openpyxl")
                    rep_data[header] = [pd.read_excel(fpath, sheet_name="Report_Regulated", usecols="F", nrows=64, engine="openpyxl").iloc[63, 0]] + df_rep.iloc[:, 0].tolist()
                except: pass
            
            max_qa, max_rep = max((len(v) for v in qa_data.values()), default=0), max((len(v) for v in rep_data.values()), default=0)
            for k in qa_data: qa_data[k].extend([None] * (max_qa - len(qa_data[k])))
            for k in rep_data: rep_data[k].extend([None] * (max_rep - len(rep_data[k])))
            
            cmb = pd.concat([pd.DataFrame(qa_data), pd.DataFrame(rep_data)], axis=1)
            if not cmb.empty:
                cmb = cmb.reindex(sorted(cmb.columns), axis=1)
                names = (["Name"] + pcb_rows)[:len(cmb)]
                cmb.insert(0, "Names", names + [None]*(len(cmb)-len(names)))
                cmb.to_excel(out_path, sheet_name="Consolidated", index=False)

    # --- Dioxin Extraction (ONLY LAST FOLDER) ---
    if dioxin_folders:
        last_dxn_folder = str(dioxin_folders[-1]).strip()
        out_path = os.path.join(last_dxn_folder, "consolidated_combined.xlsx")
        
        if os.path.exists(last_dxn_folder):
            qa_data, ids = {}, {}
            files = [f for f in os.listdir(last_dxn_folder) if f.endswith(".xlsm") and not f.startswith("~$")]
            for file in tqdm(files, desc=f"🔍 Dioxins in {os.path.basename(last_dxn_folder)}"):
                header, fpath = file[:7], os.path.join(last_dxn_folder, file)
                try:
                    df_qa = pd.read_excel(fpath, sheet_name="QA", usecols="J", skiprows=47, nrows=25, engine="openpyxl")
                    qa_data[header] = df_qa.iloc[:, 0].values.tolist()
                    ids[header] = str(pd.read_excel(fpath, sheet_name="QA", usecols="A", skiprows=44, nrows=1, engine="openpyxl").iloc[0, 0])[:6]
                except: pass
            
            max_len = max((len(v) for v in qa_data.values()), default=0)
            for k, v in qa_data.items(): qa_data[k] = v + [np.nan] * (max_len - len(v))
            
            qa_df = pd.DataFrame(qa_data)
            if ids: qa_df = pd.concat([pd.DataFrame(ids, index=["Identifier"]).T.T, qa_df], axis=0)
            if not qa_df.empty:
                c_rows = (list(dxn_rows) + [""]*len(qa_df))[:len(qa_df)]
                qa_df.index = c_rows
                qa_df.insert(0, "Names", qa_df.index)
                qa_df.loc[~qa_df["Names"].astype(str).str.strip().eq("")].to_excel(out_path, sheet_name="Consolidated", index=False)

    # --- Consolidation & Formatting (GATHERS ALL FOLDERS) ---
    def gather(folders, prefix=False):
        cmb = pd.DataFrame()
        for f in folders:
            path = os.path.join(str(f).strip(), "consolidated_combined.xlsx")
            if os.path.exists(path):
                df = pd.read_excel(path).replace("IH001", "19680") if prefix else pd.read_excel(path)
                cmb = df if cmb.empty else pd.merge(cmb, df, on="Names", how="outer")
        if not cmb.empty:
            cmb.columns = ["oec" + c if (prefix and " " not in c and c != "Names") else str(c) for c in cmb.columns]
            return cmb[["Names"] + sorted([c for c in cmb.columns if c != "Names"])]
        return cmb

    pcb_df, dxn_df = gather(pcb_folders, True), gather(dioxin_folders, False)
    if not dxn_df.empty and len(dxn_df) > 15: dxn_df = dxn_df.drop(index=15)
    
    if not pcb_df.empty or not dxn_df.empty:
        all_cols = sorted(set(pcb_df.columns).union(dxn_df.columns))
        all_cols.insert(0, all_cols.pop(all_cols.index("Names")))
        
        cmb = pd.concat([pcb_df.reindex(columns=all_cols, fill_value=0), dxn_df.reindex(columns=all_cols, fill_value=0)], ignore_index=True)
        t_df = resolve_duplicate_columns(cmb.set_index("Names").transpose()).rename_axis("Batch ID").reset_index()
        
        # Mapping Translation
        pt_map = {"im001": "19680", "IM001": "19680", "19680.0": "19680", "6133": "T06133", "6133.0": "T06133", "": "19680", "nan": "19680"}
        for col in ["Name", "Name_1"]:
            if col in t_df.columns:
                t_df[col] = t_df[col].astype(str).str.strip().str.replace(r'\.0$', '', regex=True).replace(pt_map)
                t_df = t_df[t_df[col] != "0"]
                
        fname = f"combined_PCB_Dioxins_data_{datetime.now().strftime('%Y%m%d')}.xlsx"
        save_ukas_formatted_table(t_df.reset_index(drop=True), fname, "PCBDioxinData", 15, 10)
# =====================================================================
# MODULE 2: PBDEs
# =====================================================================
def process_pbdes():
    print("\n" + "="*50 + "\n🚀 INITIATING PBDEs EXTRACTION\n" + "="*50)
    pbde_folders = get_folders_from_excel("PBDE", fallback_index=3)
    pbde_rows = ["Name", "BDE-28", "BDE-47", "BDE-49", "BDE-66", "BDE-99", "BDE-100", "BDE-153", "BDE-154", "BDE-183", "BDE-209", "", "BB-49", "BB-52", "BB-153", "BDE-119", "BDE-126", "BDE-138"]

    # --- PBDE Extraction (ONLY LAST FOLDER) ---
    if pbde_folders:
        last_pbde_folder = str(pbde_folders[-1]).strip()
        out_path = os.path.join(last_pbde_folder, "consolidated_combined.xlsx")
        
        if os.path.exists(last_pbde_folder):
            qa_data, ids, files = {}, {}, [f for f in os.listdir(last_pbde_folder) if f.endswith(".xlsm") and not f.startswith("~$")]
            for file in tqdm(files, desc=f"🔍 PBDEs in {os.path.basename(last_pbde_folder)}"):
                header, fpath = file[5:13], os.path.join(last_pbde_folder, file)
                try:
                    qa_data[header] = [0 if str(x).strip() in ('<', 'nm') else x for x in pd.read_excel(fpath, sheet_name="QA", usecols="H", skiprows=2, nrows=20, engine="openpyxl").iloc[:, 0].tolist()]
                    ids[header] = pd.read_excel(fpath, sheet_name="PBDE", header=None, engine="openpyxl").iloc[0, 13]
                except: pass
                
            max_len = max((len(v) for v in qa_data.values()), default=0)
            for k, v in qa_data.items(): qa_data[k] = v + [np.nan] * (max_len - len(v))
            
            qa_df = pd.DataFrame(qa_data)
            if ids: qa_df = pd.concat([pd.DataFrame(ids, index=["Identifier"]).T.T, qa_df], axis=0)
            if not qa_df.empty:
                c_rows = (list(pbde_rows) + [""]*len(qa_df))[:len(qa_df)]
                qa_df.index = c_rows
                qa_df.insert(0, "Names", qa_df.index)
                qa_df.loc[~qa_df["Names"].astype(str).str.strip().eq("")].to_excel(out_path, sheet_name="Consolidated", index=False)

    # Gather & Format (GATHERS ALL FOLDERS)
    cmb = pd.DataFrame()
    for f in pbde_folders:
        path = os.path.join(str(f).strip(), "consolidated_combined.xlsx")
        if os.path.exists(path):
            cmb = pd.read_excel(path) if cmb.empty else pd.merge(cmb, pd.read_excel(path), on="Names", how="outer")
            
    if not cmb.empty:
        t_df = resolve_duplicate_columns(cmb.set_index("Names").transpose()).rename_axis("Batch ID").reset_index()
        if "Name" in t_df.columns: t_df.rename(columns={"Name": "PT Round"}, inplace=True)
        
        if "PT Round" in t_df.columns:
            t_df["PT Round"] = t_df["PT Round"].astype(str).str.replace(r'\.0+$', '', regex=True).replace("nan", "")
            t_df = t_df[["Batch ID", "PT Round"] + [c for c in t_df.columns if c not in ["Batch ID", "PT Round"]]]
            t_df["PT Round"] = t_df["PT Round"].replace({"T06136": "T06133"})
            
        t_df.drop(columns=[c for c in ["BB-153", "BB-49", "BB-52"] if c in t_df.columns], inplace=True)
        if "BDE-100" in t_df.columns:
            bde100 = pd.to_numeric(t_df["BDE-100"], errors='coerce')
            t_df = t_df[bde100.notna() & (bde100 != 0)]
            
        if len(t_df.columns) > 3:
            subset = t_df.iloc[:, 3:].replace({0: np.nan, 0.0: np.nan}).copy()
            t_df.iloc[:, 3:] = subset.infer_objects(copy=False)

        fname = f"PBDEs_consolidated_{datetime.now().strftime('%Y%m%d')}.xlsx"
        save_ukas_formatted_table(t_df, fname, "PBDEData")
# =====================================================================
# UPDATED MODULE 3: PAHs
# =====================================================================
def process_pahs():
    print("\n" + "="*50 + "\n🚀 INITIATING PAHs EXTRACTION (MACRO-ENABLED)\n" + "="*50)
    pah_folders = get_folders_from_excel("PAH", fallback_index=4)
    pah_rows = ["Batch Information", "benzo[a]anthracene", "benzo[b]fluoranthene", 
                "benzo[a]pyrene", "indeno[1,2,3-cd]pyrene", "benzo[g,h,i]perylene", 
                "Chrysene", "PAH4 (sum)"]

    # --- PAH Extraction (ONLY LAST FOLDER) ---
    if pah_folders:
        last_pah_folder = str(pah_folders[-1]).strip()
        out_path = os.path.join(last_pah_folder, "consolidated_combined.xlsx")
        
        if os.path.exists(last_pah_folder):
            # DispatchEx opens a completely fresh, isolated Excel instance
            excel = win32.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.AskToUpdateLinks = False 
            
            res_data = defaultdict(list)
            files = [f for f in os.listdir(last_pah_folder) if f.endswith(".xlsm") and not f.startswith("~$")]
            
            for file in tqdm(files, desc=f"🔍 PAHs in {os.path.basename(last_pah_folder)}"):
                fpath = os.path.join(last_pah_folder, file)
                temp_fpath = fpath.replace(".xlsm", "_temp.xlsx")
                
                try:
                    wb = excel.Workbooks.Open(fpath, UpdateLinks=0, ReadOnly=True, IgnoreReadOnlyRecommended=True)
                    wb.RefreshAll()
                    wb.SaveAs(temp_fpath, FileFormat=51) 
                    wb.Close(SaveChanges=False)
                    
                    header = file[:8]
                    
                    # SAFE READ HELPERS: Prevents "index out of bounds" errors on empty sheets
                    def safe_cell(df): return df.iloc[0, 0] if not df.empty else np.nan
                    def safe_col(df): return df.iloc[:, 0].fillna(0.0).tolist() if not df.empty else []

                    # Check which format to use
                    df_b96 = pd.read_excel(temp_fpath, sheet_name="Results", usecols="B", skiprows=94, nrows=1)
                    b96_val = safe_cell(df_b96)
                    use_v2 = not (pd.isna(b96_val) or str(b96_val).strip() == "")
                    
                    # Extract Data Safely
                    if not use_v2:
                        df_vals = pd.read_excel(temp_fpath, sheet_name="Results", usecols="D", skiprows=98, nrows=7)
                        vals = safe_col(df_vals)
                        df_hdr = pd.read_excel(temp_fpath, sheet_name="Results", usecols="F", skiprows=96, nrows=1)
                        hdr = safe_cell(df_hdr)
                    else:
                        df_vals = pd.read_excel(temp_fpath, sheet_name="Results", usecols="E", skiprows=100, nrows=7)
                        vals = safe_col(df_vals)
                        df_hdr = pd.read_excel(temp_fpath, sheet_name="Results", usecols="G", skiprows=98, nrows=1)
                        hdr = safe_cell(df_hdr)
                    
                    # --- NEW DROP-OUT LOGIC ---
                    # If the header cell is empty, NaN, or explicitly N/A, skip the file entirely
                    if pd.isna(hdr) or str(hdr).strip() == "" or str(hdr).strip().upper() == "N/A" or str(hdr).strip().lower() == "nan":
                        continue
                    
                    # Pad missing values with 0.0 just in case the column had fewer than 7 rows
                    vals = (vals + [0.0]*7)[:7]
                    
                    # Since we filtered out N/A above, we can safely just append hdr
                    res_data[header].append([hdr] + vals)
                    
                except Exception as e:
                    print(f"Error processing {file}: {e}")
                finally:
                    if os.path.exists(temp_fpath):
                        try: os.remove(temp_fpath) 
                        except: pass
            
            if res_data:
                flattened = {}
                for header, data_list in res_data.items():
                    for i, entry in enumerate(data_list):
                        col_name = f"{header}_{i}" if len(data_list) > 1 else header
                        flattened[col_name] = entry
                
                df = pd.DataFrame(flattened, index=pah_rows)
                df.insert(0, "Names", pah_rows)
                df.to_excel(out_path, sheet_name="Results", index=False)

            excel.Quit() 

    # Gather & Format (GATHERS ALL FOLDERS)
    cmb = pd.DataFrame()
    for f in pah_folders:
        path = os.path.join(str(f).strip(), "consolidated_combined.xlsx")
        if os.path.exists(path):
            cmb = pd.read_excel(path) if cmb.empty else pd.merge(cmb, pd.read_excel(path), on="Names", how="outer")
            
    if not cmb.empty:
        t_df = resolve_duplicate_columns(cmb.set_index("Names").transpose()).rename_axis("Batch ID").reset_index()
        if "Batch Information" in t_df.columns: t_df.rename(columns={"Batch Information": "PT Round"}, inplace=True)
        
        if "PT Round" in t_df.columns:
            t_df["PT Round"] = t_df["PT Round"].fillna("T0658").replace({"T00658": "T0658", "": "T0658", " ": "T0658"})
            
        fname = f"PAHs_consolidated_{datetime.now().strftime('%Y%m%d')}.xlsx"
        save_ukas_formatted_table(t_df, fname, "PAHData")

# =====================================================================
# MODULE 4: PFAS / FAPAS (OPTIMIZED WITH CACHING)
# =====================================================================
def process_pfas():
    print("\n" + "="*50 + "\n🚀 INITIATING PARALLEL PFAS EXTRACTION\n" + "="*50)
    
    # 1. PULL FOLDERS & KEYWORDS AS PAIRED DICTIONARY (Col E & Col F)
    folder_targets = {}
    try:
        df_paths = pd.read_excel(PATH_LIST_FILE)
        # Iterate through Excel rows: Col E is index 4, Col F is index 5
        for idx, row in df_paths.iterrows():
            folder = str(row.iloc[4]).strip()
            keyword = str(row.iloc[5]).strip()
            
            if folder and folder.lower() != 'nan':
                if folder not in folder_targets:
                    folder_targets[folder] = []
                if keyword and keyword.lower() != 'nan':
                    folder_targets[folder].append(keyword)
        print(f"🎯 Loaded {len(folder_targets)} valid PFAS folders with their specific PT keywords.")
    except Exception as e:
        print(f"⚠️ Could not read folders/keywords from Excel: {e}")
        return

    # 2. OPTIONAL: Grab any extra generic patterns from your limits file
    global_patterns = []
    try:
        df_t = pd.read_excel(r"S:\Ops-OPH\OPHA-OPHA\Limited\OEC-OPHA\UKAS\Ref Material checks\Your_Limits_File.xlsx", sheet_name="Sheet1", engine='calamine')
        global_patterns = [p for p in df_t["PT Round"].dropna().astype(str).str.strip().unique() if p not in ["", "nan", "0", "0.0"]]
    except: pass

    cache_path = os.path.join(FINAL_OUTPUT_DIR, "pfas_cache.pkl")
    cache = {}
    if os.path.exists(cache_path):
        with open(cache_path, 'rb') as f:
            cache = pickle.load(f)

    # Helper functions for processing
    def norm(v): return re.sub(r'[^a-z0-9]', '', str(v).lower().strip()) if pd.notna(v) else ""
    def to_num(v):
        v_str = str(v).strip().lower()
        if v_str in ('', '<', 'nm', 'none', 'nd') or pd.isna(v): return 0.0
        try: return float(re.sub(r'[^\d.eE\-+]', '', v_str.replace(',', '')))
        except: return 0.0

    def process_single_file(fpath, search_patterns):
        try:
            xl = pd.ExcelFile(fpath, engine='calamine')
            pref = ["reviewed results", "reviewed", "reviewed data", "results table"]
            s_name = next((s for s in xl.sheet_names if s.lower() in pref), None)
            if not s_name: return None
            
            df_raw = pd.read_excel(fpath, sheet_name=s_name, header=None, nrows=500, engine='calamine').iloc[:, :80]
            hdr_idx = next((r for r in range(min(40, len(df_raw))) if norm("L-PFHxS") in [norm(x) for x in df_raw.iloc[r]]), 0)
            h_norms = [norm(h) for h in df_raw.iloc[hdr_idx]]
            body = df_raw.iloc[hdr_idx + 1:]
            
            for idx, val in enumerate(body.iloc[:, 1].astype(str)):
                val_lower = val.lower().strip()
                
                # CHANGED: Case-insensitive match check!
                pt_match = next((p for p in search_patterns if p.lower().strip() in val_lower), None)
                
                if pt_match:
                    row_vals = body.iloc[idx].tolist()
                    v_dict = {a: to_num(row_vals[h_norms.index(norm(a))]) if norm(a) in h_norms else 0.0 for a in ["L-PFHxS", "Br-PFHxS", "PFOA", "PFNA", "L-PFOS", "Br-PFOS"]}
                    b_match = re.search(r'(MS\d{2}-\d{5})', os.path.basename(fpath), re.I)
                    return {
                        "Batch ID": b_match.group(1).upper() if b_match else os.path.basename(fpath)[:10].strip(),
                        "PT Round": pt_match, **v_dict,
                        "Total PFHxS": v_dict["L-PFHxS"] + v_dict["Br-PFHxS"],
                        "Total PFOS": v_dict["L-PFOS"] + v_dict["Br-PFOS"],
                        "Sum of 4 PFAS": sum([v_dict[x] for x in v_dict])
                    }
        except: return None
        return None

    all_rows = []
    new_cache = cache.copy()
    
    with concurrent.futures.ThreadPoolExecutor(max_workers=8) as executor:
        for folder, specific_keywords in folder_targets.items():
            if not os.path.exists(folder): 
                print(f"⚠️ Folder not found, skipping: {folder}")
                continue
            
            current_mtime = os.path.getmtime(folder)
            if cache.get(folder) == current_mtime: 
                print(f"⏭️ Skipping (No changes detected): {os.path.basename(folder)}")
                continue
            
            # Combine the specific Col F keywords for this folder with the global limits file patterns
            combined_patterns = list(set(specific_keywords + global_patterns))
            
            files = [os.path.join(r, f) for r, d, fs in os.walk(folder) for f in fs if f.endswith(".xlsm") and not f.startswith("~$")]
            
            future_to_file = {executor.submit(process_single_file, f, combined_patterns): f for f in files}
            for future in concurrent.futures.as_completed(future_to_file):
                res = future.result()
                if res: all_rows.append(res)
            
            new_cache[folder] = current_mtime

    with open(cache_path, 'wb') as f:
        pickle.dump(new_cache, f)

    if all_rows:
        t_df = pd.DataFrame(all_rows).drop_duplicates(subset=["Batch ID"], keep="first")
        fname = f"PFAS_consolidated_{datetime.now().strftime('%Y%m%d')}.xlsx"
        save_ukas_formatted_table(t_df, fname, "PFASData")
    else:
        print("\n⚠️ No new PFAS data was extracted. (Files were either skipped by cache, or keywords didn't match inside the Excel sheets).")
        
# =====================================================================
# EXECUTION PIPELINE
# =====================================================================
if __name__ == "__main__":
    print("\n" + "*"*60)
    print("🔬 OEC MASTER POPs DATA ENGINE INITIATED")
    print("*"*60)
    
    start_time = time.time()
    
    process_pcb_dioxins()
    process_pbdes()
    process_pahs()
    process_pfas()
    
    print("\n" + "*"*60)
    print(f"🎉 MASTER PIPELINE COMPLETE IN {round(time.time() - start_time, 2)} SECONDS.")
    print("*"*60)
